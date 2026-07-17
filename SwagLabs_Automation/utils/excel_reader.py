from openpyxl import load_workbook


def read_data(file, sheet):

    workbook = load_workbook(file)

    ws = workbook[sheet]

    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        data.append(row)

    workbook.close()

    return data